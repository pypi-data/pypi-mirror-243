"""
Powered by Gabriel Lopes de Souza

For use, please, install openpyxl with pip install openpyxl
"""


import os
from time import sleep
from FuncsForSPO.fpython.functions_for_py import remove_extensao_de_str

import openpyxl


def get_names_worksheets(Workbook, print_value: bool=False) -> list[str]:
    """Retorna a(s) sheet(s) presentes()) na tabela

    Args:
        Workbook (Workbook, openpyxl): O Workbook
        print_value (bool, optional): printa o valor a ser retornado. Defaults to False.

    Returns:
        list: lista de sheets
    """
    if print_value:
        print(Workbook.sheetnames)
        return Workbook.sheetnames
    else:
        return Workbook.sheetnames

def len_columns(plan, print_value: bool=False) -> int:
    """Retorna a quantidade de colunas existentes na tabela

        pegue a planilha desse jeito:
        
        # nome da planilha (abre a planilha)
        ARQUIVO_EXCEL = openpyxl.load_workbook(os.path.abspath('ARQUIVO_EXCEL.xlsx'))
        
        nome_planilhas = ARQUIVO_EXCEL.sheetnames # saber quais são as planilhas que tem no arquivo excel
        print(nome_planilhas) # ['Página1']

        # pegar a sheet da planilha ARQUIVO_EXCEL
        PLANILHA = ARQUIVO_EXCEL['Página1']

    Args:
        plan (workbook): planilha
        print_value (bool, optional): Mostrar o valor retornado? Defaults to False.

    Returns:
        _type_: _description_
    """
    if print_value:
        print(plan.max_column)
        return plan.max_column
    else:
        return plan.max_column
    

def pega_dados_de_coluna(plan, column: int, convert_tuple: bool=True, print_values: bool=False, formatting_date: str="%d/%m/%Y") -> list[str]:
    """Retorna os dados da coluna enviada, (tem que ser o indice da coluna)

        
        pegue a planilha desse jeito:
        
        # nome da planilha (abre a planilha)
        ARQUIVO_EXCEL = openpyxl.load_workbook(os.path.abspath('ARQUIVO_EXCEL.xlsx'))
        
        nome_planilhas = ARQUIVO_EXCEL.sheetnames # saber quais são as planilhas que tem no arquivo excel
        print(nome_planilhas) # ['Página1']

        # pegar a sheet da planilha ARQUIVO_EXCEL
        PLANILHA = ARQUIVO_EXCEL['Página1']
        
    Args:
        plan (planilha): planilha feita no openpyxl
        column (int): indice da coluna na tabela
        convert_tuple (bool): converte a lista de dados para tupla
        formatting_date (str): Formatação da data em caso de celulas que sejam datetime.datetime
    Returns:
        tuple or list: _description_
    """
    dados_da_coluna = []
    for linha in plan:
        if linha[column].value is not None:  # se o valor da planilha não for none
            cell = linha[column].value
            if "datetime" in str(type(cell)):
                cell = str(cell.strftime(formatting_date))
                dados_da_coluna.append(cell)
            if cell is None:
                dados_da_coluna.append('None')
                
            else:
                dados_da_coluna.append(cell)
        else:
            dados_da_coluna.append('None')
    if print_values:
        print(f'O nome da coluna é {dados_da_coluna[0]}')  # delete o nome da coluna
        
    del dados_da_coluna[0]  # deleta o nome da coluna
    if convert_tuple:
        dados_da_coluna = tuple(dados_da_coluna)
        
    if print_values:
        print(dados_da_coluna)
        return (dados_da_coluna)
    else:
        return (dados_da_coluna)


def return_columns_names(plan, len_columns: int, convert_to_tuple: bool=True, print_values :bool=False):
    nomes_das_colunas = []
    for coluna_index in range(len_columns):
        for linha in plan:
            coluna = linha[coluna_index].value
            nomes_das_colunas.append(coluna)
            break
    if convert_to_tuple:
        nomes_das_colunas = tuple(nomes_das_colunas)
    if print_values:
        print(nomes_das_colunas)
        return nomes_das_colunas
    else:
        return nomes_das_colunas
    
# from FuncsForSPO.openpyxl_funcs import pega_dados_de_coluna

## functions_selenium ##
def volta_paginas(driver, qtd_pages_para_voltar : int=1, espera_ao_mudar : int=0) -> None:
    """
    ### Essa função volta (back) quantas páginas você desejar

    Args:
        driver (_type_): Seu webdriver
        qtd_pages_para_voltar (int): Quantidade de páginas que serão voltadas. O padrão é uma página (1).
        espera_ao_mudar (int or float, optional): Se você quer esperar um tempo para voltar uma página. O padrão é 0.
        
    Uso:
        volta_paginas(driver=self.chrome, qtd_pages_para_voltar=3, espera_ao_mudar=1)
    """
    if espera_ao_mudar == 0:
        for back in range(qtd_pages_para_voltar):
            driver.back()
            driver.refresh()
    else:
        for back in range(qtd_pages_para_voltar):
            sleep(espera_ao_mudar)
            driver.back()
            driver.refresh()
            
            
            
## functions_openpyxl ##
# verifica se a string existe naquela coluna:

def verifica_se_existe_na_coluna(a_procurar :str, dados : tuple or list, return_list_de_achados:bool) -> bool:
    lista_de_achados = []
    for dado in dados:
        if dado == a_procurar:
            lista_de_achados.append(dado)
            print(f'Achou o dado -> {dado}') 
    else:
        if len(lista_de_achados) == 0:
            print('Não achou nenhum dado!')
        else:
            if return_list_de_achados:
                print(f'Achou {len(lista_de_achados)} item(ns)')
                return lista_de_achados
            else:
                print(f'Achou {len(lista_de_achados)} item(ns)')
                return lista_de_achados
            
            

arquivo = 'arq.xlsx'

def cria_planilha(sheets : tuple, nome_da_planilha : str='Table', create_file :bool=True) -> None:
    """Cria uma planilha 
    ### (CASO ELA JÁ EXISTA, SERÁ SUBSTITUIDA!)
    
    ## Atenção as sheets deverão ser enviadas desse modo:
    #### sheets = (('sheet1', 0), ('sheet2', 1),)
    
    #### sheets = (('sheet1', 0), ('sheet2', 1), ('sheet3', -1),)
    
    As sheets serão desempacotadas na função create_sheet('sheet1', 0)

    Args:
        nome_da_planilha (str): Nome da planilha que ficará no arquivo
        sheets (tuple): sheets da planilha que deve vir desse jeito -> (('sheet1', 0), ('sheet2', 1), ('sheet3', -1),)
    """
    nome_da_planilha = remove_extensao_de_str(nome_da_planilha, 'xlsx')

    wb = openpyxl.Workbook()  # cria um arquivo
    del wb['Sheet']  # deleta a planilha criada automaticamente

    for sheet in sheets:
        ws = wb.create_sheet(*sheet)
        print(f'A WorkSheet "{ws.title}" foi criada!')
    if create_file:
        print(f'\nO arquivo {nome_da_planilha}.xlsx foi criado!')
        wb.save(nome_da_planilha+'.xlsx')
    else:
        return wb
    
    
def recupera_worksheet_do_arquivo(file_plan, ws_a_recuperar):
    try:
        wb = openpyxl.load_workbook(os.path.abspath(file_plan))
    except FileNotFoundError:
        print('Não existe o arquivo para recuperar a worksheet')
        quit()
    try:
        WORKSHEET = wb[ws_a_recuperar]
    except KeyError:
        print(f'Não existe uma Worksheet com o nome {ws_a_recuperar} na tabela! (FUNÇÃO {recupera_worksheet_do_arquivo.__name__})')
        return
    return WORKSHEET
    
def deleta_planilha_da_tabela(Plan, sheet_name_for_del) -> None:
    plan = recupera_worksheet_do_arquivo()
    del sheet_name_for_del

def cria_colunas_na_worksheet(plan_file, worksheet, cols : tuple):
    wb = openpyxl.load_workbook(os.path.abspath(plan_file))
    try:
        ws = wb[worksheet]
    except KeyError:
        print(f'Não existe uma Worksheet com o nome {worksheet} na tabela! (FUNÇÃO {cria_colunas_na_worksheet.__name__})')
        return
    for col in cols:
        ws[col[0]] = col[1]

    
def pega_nome_das_colunas(plan_file, sheet):
    wb = openpyxl.load_workbook(os.path.abspath(plan_file))
    try:
        ws = wb[sheet]
    except KeyError:
        print(f'Não existe uma Worksheet com o nome {sheet} na tabela! (FUNÇÃO {pega_nome_das_colunas.__name__})')
        return
    list_with_values=[]
    for cell in ws[1]:
        list_with_values.append(cell.value)
    return list_with_values

def adiciona_dados_em_uma_coluna(plan_file, sheet, col :int, your_list :list, row_start=2):
    """col é o indice da coluna"""
    wb = openpyxl.load_workbook(os.path.abspath(plan_file))
    ws = wb[sheet]
    
    for i, value in enumerate(your_list, start=row_start):
        ws.cell(row=i, column=col).value = value

    # wb.save(plan_file)

# adiciona_dados_em_uma_coluna('Table.xlsx', 'sheet1', 2, ['João', 'Gabriel', 'Irineu', 'malandro', 'severino'],)
# linhas = 
def criador_de_arquvos_xlsx(nome_do_arquivo : str, sheet : tuple, colunas : tuple, dados : tuple, linha_para_adicionar_os_dados=2) -> None:
    """Cria um arquivo xlsx com openpyxl

    Args:
        nome_do_arquivo (str): Nome do arquivo SEM A EXTENSÃO
        
        sheet (tuple): SHEET QUE DEVE VIR NESSE FORMATO -> ('Resultado', 0)
        
        colunas (tuple): colunas que DEVE VIR NESSE FORMATO -> (('A1', 'NOME'), ('B1', 'SOBRENOME'), ('C1', 'IDADE'))
        
        dados (tuple):  dado que DEVE VIR NESSE FORMATO (COM O MESMO LEN DAS COLUNAS!) -> ((1, [1,2,3,4,5,11]), (2, [1,2,3,4,5]), (3, [1,2,3,4,5]))

        linha_para_adicionar_os_dados (int, optional): LINHA QUE DEVE COMEÇAR A ENVIAR OS DADOS!. Defaults to 2 POIS ADICIONARÁ SEMPRE UMA NOVA.

    """
    # cria+planilha
    SHEET_NAME = sheet[0]

    if len(colunas) == len(dados):    
        wb = openpyxl.Workbook()  # cria um arquivo
        del wb['Sheet']  # deleta a planilha criada automaticamente

        ws = wb.create_sheet(*sheet)  # Desempacota os dados da sheet
        print(f'A WorkSheet "{ws.title}" foi criada!')

        try:
            ws = wb[SHEET_NAME]
        except KeyError:
            print(f'Não existe uma Worksheet com o nome {SHEET_NAME} na tabela!')
            return
        
        for col in colunas:
            ws[col[0]] = col[1]
            
        for t in dados:
            col = t[0]
            for i, value in enumerate(t[1], start=linha_para_adicionar_os_dados):
                ws.cell(row=i, column=col).value = value
        wb.save(os.path.abspath(f"{nome_do_arquivo}.xlsx"))
    else:
        print(f'Existe itens a mais ou a menos existem {len(colunas)} colunas e {len(dados)} dados a serem enviados')
    
if __name__ == '__main__':
    dados = ((1, [1,2,3,4,5,11]), (2, [1,2,3,4,5]), (3, [1,2,3,4,5]))
    cols = (('A1', 'NOME'), ('B1', 'SOBRENOME'), ('C1', 'IDADE'))
    sheet = ('Resultado', 0)
    ws = recupera_worksheet_do_arquivo('MeuArquivo.xlsx', 'Resultado')
    dados_verifica = pega_dados_de_coluna(plan=ws, column=0, print_values=True)
    criador_de_arquvos_xlsx(nome_do_arquivo='MeuArquivo', sheet=sheet, colunas=cols, dados=dados)
    verifica_se_existe_na_coluna(a_procurar='10', dados=dados_verifica, return_list_de_achados=True)