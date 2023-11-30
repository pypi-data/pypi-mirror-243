#!/usr/bin/env python3
# -*- coding: utf-8 -*-
""" Module to allow user to make changes to Assessments in an excel spreadsheet for user friendly experience """

# standard python imports
import math
import os
import shutil
from pathlib import Path
from typing import Union

import click
import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection, Font, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation

from regscale.core.app.api import Api
from regscale.core.app.application import Application
from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import (
    check_file_path,
    error_and_exit,
    reformat_str_date,
    get_user_names,
    get_current_datetime,
    check_empty_nan,
)
from regscale.models.app_models.click import regscale_id, regscale_module
from regscale.models.regscale_models.assessment import Assessment
from regscale.models.regscale_models.modules import Modules


@click.group(name="assessments")
def assessments():
    """
    Performs actions on Assessments CLI Feature to create new or update assessments to RegScale.
    """


# Make Empty Spreadsheet for creating new assessments.
@assessments.command(name="generate_new_file")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path for excel files to be generated into.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def generate_new_file(path: Path):
    """This function will build an excel spreadsheet for users to be able to create new assessments."""
    new_assessment(path)


def new_assessment(path: Path) -> None:
    """Function to build excel spreadsheet for creation of new assessments.

    :param Path path: directory of file location
    :return: None
    """
    logger = create_logger()

    check_file_path(path)

    # create excel file and setting formatting

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "New_Assessments"

    column_headers = [
        "Title",
        "LeadAssessor",
        "Facility",
        "Organization",
        "AssessmentType",
        "PlannedStart",
        "PlannedFinish",
        "Status",
        "ActualFinish",
        "AssessmentResult",
        "ParentId",
        "ParentModule",
    ]
    for col, val in enumerate(column_headers, start=1):
        worksheet.cell(row=1, column=col).value = val

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        for cell in worksheet[col]:
            if cell.row == 1:
                cell.font = Font(bold=True)

    # create and format reference worksheets for dropdowns
    workbook.create_sheet(title="Facilities")
    workbook.create_sheet(title="Organizations")
    workbook.create_sheet(title="Accounts")
    workbook.create_sheet(title="Modules")
    workbook.create_sheet(title="AssessmentTypes")
    workbook.create_sheet(title="Assessment_Ids")

    workbook.save(filename=os.path.join(path, "new_assessments.xlsx"))

    # pull in Facility, Organization, Module, and Account Usernames into Excel Spreadsheet to create drop downs
    list_of_modules = Modules().api_names()
    module_names = pd.DataFrame(list_of_modules, columns=["name"])
    with pd.ExcelWriter(
        os.path.join(path, "new_assessments.xlsx"),
        mode="a",
        engine="openpyxl",
        if_sheet_exists="overlay",
    ) as writer:
        get_field_names(field_name="facilities").to_excel(
            writer,
            sheet_name="Facilities",
            index=False,
        )
        get_field_names(field_name="organizations").to_excel(
            writer,
            sheet_name="Organizations",
            index=False,
        )
        get_user_names().to_excel(
            writer,
            sheet_name="Accounts",
            index=False,
        )
        module_names.to_excel(
            writer,
            sheet_name="Modules",
            index=False,
        )
        get_assessment_types().to_excel(
            writer,
            sheet_name="AssessmentTypes",
            index=False,
        )

    # Creating data Validation for fields
    workbook = load_workbook(os.path.join(path, "new_assessments.xlsx"))
    worksheet = workbook.active
    facilities_worksheet = workbook["Facilities"]
    accounts_worksheet = workbook["Accounts"]
    organizations_worksheet = workbook["Organizations"]
    assessment_worksheet = workbook["AssessmentTypes"]
    modules_worksheet = workbook["Modules"]

    # lock worksheets containing data for dropdowns
    facilities_worksheet.protection.sheet = True
    accounts_worksheet.protection.sheet = True
    organizations_worksheet.protection.sheet = True
    assessment_worksheet.protection.sheet = True
    modules_worksheet.protection.sheet = True
    dv1 = DataValidation(
        type="list",
        formula1="=Accounts!$A$2:$A$"
        + str(get_maximum_rows(sheet_object=workbook["Accounts"])),
        allow_blank=False,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv2 = DataValidation(
        type="list",
        formula1="=Facilities!$A$2:$A$"
        + str(get_maximum_rows(sheet_object=workbook["Facilities"])),
        allow_blank=True,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv3 = DataValidation(
        type="list",
        formula1="=Organizations!$A$2:$A$"
        + str(get_maximum_rows(sheet_object=workbook["Organizations"])),
        allow_blank=True,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv4 = DataValidation(
        type="list",
        formula1="=AssessmentTypes!$A$2:$A$"
        + str(get_maximum_rows(sheet_object=workbook["AssessmentTypes"])),
        allow_blank=False,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv5 = DataValidation(
        type="list",
        formula1='"Scheduled, In Progress, Complete, Cancelled"',
        allow_blank=True,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv6 = DataValidation(
        type="list",
        formula1='"Pass, Fail, N/A, Partial Pass"',
        allow_blank=True,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv7 = DataValidation(
        type="date",
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        showInputMessage=True,
        error="Your entry is not a valid option",
        errorTitle="Invalid Entry",
        prompt="Please enter valid date mm/dd/yyyy",
    )
    dv8 = DataValidation(
        type="list",
        formula1="=Modules!$A$2:$A$"
        + str(get_maximum_rows(sheet_object=workbook["Modules"])),
        allow_blank=True,
        showDropDown=False,
        error="Your entry is not a valid option",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv9 = DataValidation(
        type="date",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        showInputMessage=True,
        error="Your entry is not a valid option",
        errorTitle="Invalid Entry",
        prompt="Please enter valid date mm/dd/yyyy",
    )
    worksheet.add_data_validation(dv1)
    worksheet.add_data_validation(dv2)
    worksheet.add_data_validation(dv3)
    worksheet.add_data_validation(dv4)
    worksheet.add_data_validation(dv5)
    worksheet.add_data_validation(dv6)
    worksheet.add_data_validation(dv7)
    worksheet.add_data_validation(dv8)
    worksheet.add_data_validation(dv9)
    dv1.add("B2:B1048576")
    dv2.add("C2:C1048576")
    dv3.add("D2:D1048576")
    dv4.add("E2:E1048576")
    dv5.add("H2:H1048576")
    dv6.add("J2:J1048576")
    dv7.add("F2:F1048576")
    dv7.add("G2:G1048576")
    dv9.add("I2:I1048576")
    dv8.add("L2:L1048576")

    workbook.save(filename=os.path.join(path, "new_assessments.xlsx"))

    # Freezing top row and adding data style to date columns to assure validation

    workbook = load_workbook(os.path.join(path, "new_assessments.xlsx"))
    worksheet = workbook.active
    freeze_range = worksheet.cell(2, 14)
    worksheet.freeze_panes = freeze_range
    date_style = NamedStyle(name="date_style", number_format="mm/dd/yyyy")
    workbook.add_named_style(date_style)

    for col in ["F", "G", "I"]:  # Columns to edit
        for cell in worksheet[col]:
            if cell.row > 1:
                cell.style = date_style

    # Adjusting width of columns

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(filename=os.path.join(path, "new_assessments.xlsx"))

    logger.info(
        "Your excel workbook has been created. Please open the new_assessments workbook and add new assessments."
    )
    return None


@assessments.command(name="generate")
@regscale_id()
@regscale_module()
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path for excel files to be generated into.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def generate(regscale_id: str, regscale_module: str, path: Path):
    """This function will build and populate a spreadsheet of all assessments
    with the selected RegScale Parent Id and RegScale Module for users to any necessary edits.
    """
    all_assessments(regscale_id=regscale_id, regscale_module=regscale_module, path=path)


def all_assessments(regscale_id: str, regscale_module: str, path: Path) -> None:
    """Function takes organizer record and module and build excel worksheet of assessments

    :param str regscale_id: RegScale Parent Id
    :param str regscale_module: RegScale Parent Module
    :param Path path: directory of file location
    :return: None
    """
    logger = create_logger()
    app = Application()
    api = Api(app)

    body = """
            query {
                  assessments (skip: 0, take: 50, where: {parentId: {eq: parent_id} parentModule: {eq: "parent_module"}}) {
                    items {
                      id
                      title
                      leadAssessor {
                        firstName
                        lastName
                        userName
                      }
                      facility {
                        name
                      }
                      org {
                        name
                      }
                      assessmentType
                      plannedStart
                      plannedFinish
                      status
                      actualFinish
                      assessmentResult
                      parentId
                      parentModule
                    }
                    totalCount
                    pageInfo {
                      hasNextPage
                    }
                }
            }
                """.replace(
        "parent_module", regscale_module
    ).replace(
        "parent_id", str(regscale_id)
    )
    existing_assessment_data = api.graph(query=body)

    if (
        existing_assessment_data["assessments"]["totalCount"] > 0
    ):  # Checking to see if assessment exists for selected RegScale Id and RegScale Module.
        check_file_path(path)

        # Loading data from db into two workbooks.
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Assessments({regscale_id}_{regscale_module})"
        workbook.create_sheet(title="Facilities")
        workbook.create_sheet(title="Organizations")
        workbook.create_sheet(title="Accounts")
        workbook.create_sheet(title="AssessmentTypes")
        workbook.save(filename=os.path.join(path, "all_assessments.xlsx"))
        shutil.copy(
            os.path.join(path, "all_assessments.xlsx"),
            os.path.join(path, "old_assessments.xlsx"),
        )

        raw_data = existing_assessment_data["assessments"]["items"]
        assessments_data = []
        for a in raw_data:
            Id = a["id"]
            Title = a["title"]
            LeadAssessor = (
                str(a["leadAssessor"]["lastName"]).strip()
                + ", "
                + str(a["leadAssessor"]["firstName"]).strip()
                + " ("
                + str(a["leadAssessor"]["userName"]).strip()
                + ")"
            )
            Facility = a["facility"]["name"] if a["facility"] else None
            Organization = a["org"]["name"] if a["org"] else None
            AssessmentType = a["assessmentType"]
            PlannedStart = reformat_str_date(a["plannedStart"])
            PlannedFinish = reformat_str_date(a["plannedFinish"])
            Status = a["status"]
            ActualFinish = (
                reformat_str_date(a["actualFinish"]) if a["actualFinish"] else None
            )
            AssessmentResult = a["assessmentResult"] or None
            ParentId = a["parentId"]
            ParentModule = a["parentModule"]

            assessments_data.append(
                [
                    Id,
                    Title,
                    LeadAssessor,
                    Facility,
                    Organization,
                    AssessmentType,
                    PlannedStart,
                    PlannedFinish,
                    Status,
                    ActualFinish,
                    AssessmentResult,
                    ParentId,
                    ParentModule,
                ]
            )

        all_ass_df = pd.DataFrame(
            assessments_data,
            columns=[
                "Id",
                "Title",
                "LeadAssessor",
                "Facility",
                "Organization",
                "AssessmentType",
                "PlannedStart",
                "PlannedFinish",
                "Status",
                "ActualFinish",
                "AssessmentResult",
                "ParentId",
                "ParentModule",
            ],
        )

        with pd.ExcelWriter(
            os.path.join(path, "all_assessments.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_ass_df.to_excel(
                writer,
                sheet_name=f"Assessments({regscale_id}_{regscale_module})",
                index=False,
            )
        with pd.ExcelWriter(
            os.path.join(path, "old_assessments.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_ass_df.to_excel(
                writer,
                sheet_name=f"Assessments({regscale_id}_{regscale_module})",
                index=False,
            )

        # Pulling in Facility Names into Excel Spreadsheet to create dropdown.
        with pd.ExcelWriter(
            os.path.join(path, "all_assessments.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            get_field_names(field_name="facilities").to_excel(
                writer,
                sheet_name="Facilities",
                index=False,
            )
            get_field_names(field_name="organizations").to_excel(
                writer,
                sheet_name="Organizations",
                index=False,
            )
            get_user_names().to_excel(
                writer,
                sheet_name="Accounts",
                index=False,
            )
            get_assessment_types().to_excel(
                writer,
                sheet_name="AssessmentTypes",
                index=False,
            )

        # Adding protection to "old_assessments.xlsx" file that will be used as reference.
        workbook2 = load_workbook(os.path.join(path, "old_assessments.xlsx"))
        worksheet2 = workbook2.active
        worksheet2.protection.sheet = True
        workbook2.save(filename=os.path.join(path, "old_assessments.xlsx"))

        # Adding Data Validation to "all_assessments.xlsx" file to be adjusted internally.
        workbook = load_workbook(os.path.join(path, "all_assessments.xlsx"))
        worksheet = workbook.active
        facilities_worksheet = workbook["Facilities"]
        accounts_worksheet = workbook["Accounts"]
        organizations_worksheet = workbook["Organizations"]
        assessments_worksheet = workbook["AssessmentTypes"]

        # lock worksheets containing data for dropdowns
        facilities_worksheet.protection.sheet = True
        accounts_worksheet.protection.sheet = True
        organizations_worksheet.protection.sheet = True
        assessments_worksheet.protection.sheet = True
        worksheet.protection.sheet = True

        dv1 = DataValidation(
            type="list",
            formula1="=Accounts!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["Accounts"])),
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv2 = DataValidation(
            type="list",
            formula1="=Facilities!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["Facilities"])),
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv3 = DataValidation(
            type="list",
            formula1="=Organizations!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["Organizations"])),
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv4 = DataValidation(
            type="list",
            formula1="=AssessmentTypes!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["AssessmentTypes"])),
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv5 = DataValidation(
            type="list",
            formula1='"Scheduled, In Progress, Complete, Cancelled"',
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv6 = DataValidation(
            type="list",
            formula1='"Pass, Fail, N/A, Partial Pass"',
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv7 = DataValidation(
            type="date",
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not a valid option",
            showErrorMessage=True,
            showInputMessage=True,
            errorTitle="Invalid Entry",
            prompt="Please enter valid date mm/dd/yyyy",
        )
        dv8 = DataValidation(
            type="date",
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not a valid option",
            showErrorMessage=True,
            showInputMessage=True,
            errorTitle="Invalid Entry",
            prompt="Please enter valid date mm/dd/yyyy",
        )
        worksheet.add_data_validation(dv1)
        worksheet.add_data_validation(dv2)
        worksheet.add_data_validation(dv3)
        worksheet.add_data_validation(dv4)
        worksheet.add_data_validation(dv5)
        worksheet.add_data_validation(dv6)
        worksheet.add_data_validation(dv7)
        worksheet.add_data_validation(dv8)
        dv1.add("C2:C1048576")
        dv2.add("D2:D1048576")
        dv3.add("E2:E1048576")
        dv4.add("F2:F1048576")
        dv5.add("I2:I1048576")
        dv6.add("K2:K1048576")
        dv7.add("G2:G1048576")
        dv7.add("H2:H1048576")
        dv8.add("J2:J1048576")

        # Worksheet freeze top row
        freeze_range = worksheet.cell(2, 17)
        worksheet.freeze_panes = freeze_range
        date_style = NamedStyle(name="date_style", number_format="mm/dd/yyyy")
        workbook.add_named_style(date_style)

        # Adding Date Style to Worksheet, formatting cells, and unlocking cells that can be edited in each assessment

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

            if col == ["C", "D", "E", "F", "G", "H", "I", "J", "K"]:  # Columns to edit
                for cell in worksheet[col]:
                    cell.protection = Protection(locked=False)
                if col == ["G", "H", "J"]:
                    for cell in worksheet[col]:
                        if cell.row > 1:
                            cell.style = date_style

        workbook.save(filename=os.path.join(path, "all_assessments.xlsx"))

    else:
        logger.info(
            "Please check your selections for RegScale Id and RegScale Module and try again."
        )
        error_and_exit(
            "There was an error creating your workbook. No assessments exist for the given RegScale Id and RegScale Module."
        )

    return logger.info(
        "Your data has been loaded into your excel workbook. Please open the all_assessments workbook and make your desired changes."
    )


@assessments.command(name="load")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of excel workbook locations.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def load(path: Path) -> None:
    """
    This function uploads updated assessments and new assessments to the RegScale from the Excel files that users have edited.
    """
    upload_data(path=path)


def upload_data(path: Path) -> None:
    """Function will upload assessments to RegScale if user as made edits to any of the assessment excel workbooks.

    :param Path path: directory of file location
    :return: None
    """
    logger = create_logger()
    app = Application()
    config = app.config
    api = Api(app)

    # Checking if new assessments have been created and updating RegScale database.
    if os.path.isfile(os.path.join(path, "new_assessments.xlsx")):
        new_files = os.path.join(path, "new_assessments.xlsx")
        new = pd.read_excel(new_files)
        new["Facility"] = new["Facility"].fillna("None")
        new["Organization"] = new["Organization"].fillna("None")
        facilities = pd.read_excel(new_files, sheet_name="Facilities")
        facilities = facilities.rename(columns={"name": "Facility", "id": "FacilityId"})
        organizations = pd.read_excel(new_files, sheet_name="Organizations")
        organizations = organizations.rename(
            columns={"name": "Organization", "id": "OrganizationId"}
        )
        accounts = pd.read_excel(new_files, sheet_name="Accounts")
        accounts = accounts.rename(
            columns={"User": "LeadAssessor", "UserId": "LeadAssessorId"}
        )
        new = new.merge(accounts, how="left", on="LeadAssessor")
        new = new.merge(facilities, how="left", on="Facility")
        new = new.merge(organizations, how="left", on="Organization")
        new = new.T.to_dict()
        new_assessments = [
            Assessment(
                leadAssessorId=value["LeadAssessorId"],
                title=value["Title"],
                assessmentType=value["AssessmentType"],
                plannedStart=value["PlannedStart"],
                plannedFinish=value["PlannedFinish"],
                status=value["Status"],
                parentModule=check_empty_nan(value["ParentModule"]),
                facilityId=check_empty_nan(value["FacilityId"]),
                orgId=check_empty_nan(value["OrganizationId"]),
                assessmentResult=check_assessment_result(value["AssessmentResult"]),
                actualFinish=check_empty_nan(value["ActualFinish"]),
                parentId=check_empty_nan(value["ParentId"]),
                lastUpdatedById=app.config["userId"],
                dateLastUpdated=get_current_datetime(),
            ).dict()
            for value in new.values()
        ]

        try:
            for i in new_assessments:
                response = api.post(
                    url=config["domain"] + "/api/assessments",
                    json=i,
                )
                if response.ok:
                    ids = []
                    for i in response:
                        assessment_ids = response.json()["id"]
                    ids.append(assessment_ids)
                    new_assessments_df = pd.DataFrame(ids, columns=["id_number"])

                    for file_name in ["new_assessments.xlsx", "all_assessments.xlsx"]:
                        with pd.ExcelWriter(
                            os.path.join(path, file_name),
                            mode="a",
                            engine="openpyxl",
                            if_sheet_exists="overlay",
                        ) as writer:
                            new_assessments_df.to_excel(
                                writer,
                                sheet_name="Assessment_Ids",
                                index=False,
                            )

                    logger.info(
                        "%s total assessments were added to RegScale.",
                        str(len(new_assessments)),
                    )
                else:
                    logger.warning("Unable to post new assessments to RegScale")
        except requests.exceptions.RequestException as ex:
            logger.error(
                "Unable to add %i assessment(s) to RegScale.\n%s",
                len(new_assessments),
                ex,
            )
    else:
        logger.info("No new assessments detected. Checking for edited assessments")

    if os.path.isfile(os.path.join(path, "all_assessments.xlsx")):
        # Checking all_assessments file for differences before updating database

        df1 = pd.read_excel(
            os.path.join(path, "old_assessments.xlsx"), sheet_name=0, index_col="Id"
        )

        df2 = pd.read_excel(
            os.path.join(path, "all_assessments.xlsx"), sheet_name=0, index_col="Id"
        )

        if df1.equals(df2):
            error_and_exit("No differences detected.")

        else:
            logger.warning("Differences found!")

        diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
        ne_stacked = diff_mask.stack()
        changed = ne_stacked[ne_stacked]
        changed.index.names = ["Id", "Column"]
        difference_locations = np.where(diff_mask)
        changed_from = df1.values[difference_locations]
        changed_to = df2.values[difference_locations]
        changes = pd.DataFrame(
            {"From": changed_from, "To": changed_to}, index=changed.index
        )
        changes.to_csv(
            os.path.join(path, "differences.txt"),
            header=True,
            index=True,
            sep=" ",
            mode="w+",
        )
        logger.info(
            "Please check differences.txt file located in %s to see changes made.",
            path,
        )
        # Loading in differences.txt file and using Id to parse xlsx file for rows to update

        diff = pd.read_csv(
            os.path.join(path, "differences.txt"), header=0, sep=" ", index_col=None
        )
        ids = []

        for i, row in diff.iterrows():
            ids.append(row["Id"])

        id_df = pd.DataFrame(ids, index=None, columns=["Id"])
        id_df2 = id_df.drop_duplicates()
        updated_files = os.path.join(path, "all_assessments.xlsx")
        df3 = pd.read_excel(updated_files, sheet_name=0, index_col=None)
        updated = df3[df3["Id"].isin(id_df2["Id"])]
        updated["Facility"] = updated["Facility"].fillna("None")
        updated["Organization"] = updated["Organization"].fillna("None")
        facilities = pd.read_excel(updated_files, sheet_name="Facilities")
        facilities = facilities.rename(columns={"name": "Facility", "id": "FacilityId"})
        organizations = pd.read_excel(updated_files, sheet_name="Organizations")
        organizations = organizations.rename(
            columns={"name": "Organization", "id": "OrganizationId"}
        )
        accounts = pd.read_excel(updated_files, sheet_name="Accounts")
        accounts = accounts.rename(
            columns={"User": "LeadAssessor", "UserId": "LeadAssessorId"}
        )
        updated = updated.merge(accounts, how="left", on="LeadAssessor")
        updated = updated.merge(facilities, how="left", on="Facility")
        updated = updated.merge(organizations, how="left", on="Organization")
        updated = updated.T.to_dict()
        updated_assessments = [
            Assessment(
                leadAssessorId=value["LeadAssessorId"],
                id=value["Id"],
                title=value["Title"],
                assessmentType=value["AssessmentType"],
                plannedStart=value["PlannedStart"],
                plannedFinish=value["PlannedFinish"],
                status=value["Status"],
                parentModule=value["ParentModule"],
                facilityId=check_empty_nan(value["FacilityId"]),
                orgId=check_empty_nan(value["OrganizationId"]),
                assessmentResult=check_assessment_result(value["AssessmentResult"]),
                actualFinish=check_empty_nan(value["ActualFinish"]),
                parentId=value["ParentId"],
                lastUpdatedById=app.config["userId"],
                dateLastUpdated=get_current_datetime(),
            ).dict()
            for value in updated.values()
        ]

        api.update_server(
            url=config["domain"] + "/api/assessments",
            json_list=updated_assessments,
            message="Working on uploading updated assessments to RegScale.",
            config=config,
            method="put",
        )

    else:
        logger.info("No Assessments exist to load to RegScale.")
    return logger.info(
        "Assessment files have been uploaded. Changes made to existing files can be seen in differences.txt file. Thank you!"
    )


@assessments.command(name="delete_files")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of file location.",
    default=Path("./artifacts"),
    required=True,
)
def generate_delete_file(path: Path):
    """This command will delete files used during the Assessment editing process."""
    delete_file(path)


def delete_file(path: Path) -> int:
    """
    Deletes files used during the process

    :param Path path: directory of file location
    :return: Number of files deleted
    :rtype: int
    """
    logger = create_logger()
    file_names = [
        "new_assessments.xlsx",
        "all_assessments.xlsx",
        "old_assessments.xlsx",
        "differences.txt",
    ]
    deleted_files = []

    for file_name in file_names:
        if os.path.isfile(path / file_name):
            os.remove(path / file_name)
            deleted_files.append(file_name)
        else:
            logger.warning(
                "No %s file found. Checking for other files before exiting.", file_name
            )
    logger.info(
        "%i files have been deleted: %s", len(deleted_files), ", ".join(deleted_files)
    )
    return len(deleted_files)


def get_maximum_rows(*, sheet_object):
    """This function finds the last row containing data in a spreadsheet

    :param sheet_object: excel worksheet to be referenced
    :return: integer representing last row with data in spreadsheet
    :rtype: int
    """
    return sum(
        any(col.value is not None for col in row)
        for max_row, row in enumerate(sheet_object, 1)
    )


def get_field_names(field_name) -> pd.DataFrame:
    """
    This function uses GraphQL to retrieve all names of a given parent table in database

    :return: pandas dataframe with facility names
    :rtype: pd.dataframe
    """
    app = Application()
    api = Api(app)

    body = """
    query {
        field_name(skip: 0, take: 50, order: {name: ASC}, ) {
            items {
                name
                id
            }
            totalCount
            pageInfo {
                hasNextPage
            }
        }
    }
    """.replace(
        "field_name", field_name
    )

    field_items = api.graph(query=body)
    names = field_items[str(field_name)]["items"]
    field_names = [[i["name"], i["id"]] for i in names]
    all_names = pd.DataFrame(field_names, index=None, columns=["name", "id"])

    return all_names


def get_assessment_types() -> pd.DataFrame:
    """This function uses GraphQL to retrieve all assessment types in database

    :return: pandas dataframe with assessment types
    :rtype: pd.dataframe"""
    app = Application()
    api = Api(app)

    body = """
        query{
          assessments (skip: 0, take: 50, order: {assessmentType: ASC}, ) {
            items {
              assessmentType
            }
            totalCount
            pageInfo {
              hasNextPage
            }
          }
        } """

    assessments_raw = api.graph(query=body)
    assessmentTypes = assessments_raw["assessments"]["items"]
    field_names = [i["assessmentType"] for i in assessmentTypes]
    all_assessmentTypes = pd.DataFrame(
        field_names, index=None, columns=["assessmentType"]
    )
    all_assessmentTypes = all_assessmentTypes.drop_duplicates()

    return all_assessmentTypes


def check_assessment_result(value) -> Union[str, float]:
    """This function takes a given value for an assessment and checks if value is empty or NaN based on value type.

    :param value: A string or float object
    :return: A string value, float value. or ""
    :rtype: Union[str, float]
    """
    # this function has to be checked separate to account for API only accepting empty string unlike other class params
    if isinstance(value, str) and value.strip() == "":
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return value
