from collections import defaultdict
from pathlib import Path

import click
import requests
from click.exceptions import ClickException
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, Border, Side, NamedStyle

RAUM_LZ_COLUMN = "F"
RAUM_DATA_STARTS_AT_ROW = 4
FTE_DATA_STARTS_AT_ROW = 2
lz_institute = {}
instlz_name = {}
inst_rtype_area = defaultdict(lambda: defaultdict(int))
rtypes = []
summary = defaultdict(lambda: defaultdict(int))
institute_fte = defaultdict(int)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
title_style = NamedStyle("title_style")
title_style.border = thin_border
title_style.font = Font(bold=True)


@click.command()
@click.argument("raum-liste", type=click.Path(exists=True))
@click.argument("fte-liste", type=click.Path(exists=True))
@click.version_option(prog_name="Raum Management Tool")
def cli(raum_liste, fte_liste):
    """Raum Management Tool"""
    raum_liste = Path(raum_liste)
    raum_wb = load_workbook(filename=raum_liste)
    raum_wb_copy = raum_liste.parent / (
        raum_liste.stem + "_mit_Institut" + raum_liste.suffix
    )
    create_raum_wb_with_institute(raum_wb, raum_wb_copy)

    lz_institute_filepath = raum_liste.parent / "lz_instute_mapping.xlsx"
    create_lz_institute(lz_institute_filepath)

    fte_wb = load_workbook(filename=fte_liste)
    summary_filepath = raum_liste.parent / "summary_area_per_institute.xlsx"
    create_summary(fte_wb, summary_filepath)


def create_lz_institute(lz_institute_filepath):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Gruppe LZ", "Gruppe", "Institut LZ", "Institut"])
    for cell in sheet["1"]:
        cell.style = title_style
    for k, v in lz_institute.items():
        sheet.append(
            [
                k,
                v.get("oe_name"),
                v.get("oe_leitzahl_oben"),
                v.get("oe_name_oben"),
            ]
        )

    workbook.save(lz_institute_filepath)


def create_raum_wb_with_institute(raum_wb: Workbook, raum_wb_copy):
    """Copies the given Raumliste into a new Excel file
    and adds Institut-Leitzahl plus a few useful deep-links
    to ETH-internal websites
    """
    sheet = raum_wb["IM_041_Raumliste_mit_Zuteilung_"]
    raum_wb.add_named_style(title_style)

    sheet["K3"] = "Institut LZ"
    sheet["K3"].style = title_style
    sheet["L3"] = "Institut"
    sheet["L3"].style = title_style

    for row in sheet.iter_rows(min_row=RAUM_DATA_STARTS_AT_ROW):
        lz_cell = row[5]

        # Institute und Leitzahl-Hyperlinks
        inst = get_institute_for_lz(lz_cell.value)
        inst_lz = str(inst.get("oe_leitzahl_oben"))
        sheet[f"K{lz_cell.row}"] = inst_lz
        sheet[
            f"K{lz_cell.row}"
        ].hyperlink = (
            f"https://www.bi.id.ethz.ch/orgdb/OrgeinheitDetailPre.do?leitzahl={inst_lz}"
        )
        sheet[f"K{lz_cell.row}"].data_type = "s"
        sheet[f"K{lz_cell.row}"].style = "Hyperlink"
        sheet[f"L{lz_cell.row}"] = inst.get("oe_name_oben")

        # Leitzahl-Hyperlink
        lz_str = "0" * (5 - len(str(lz_cell.value))) + str(lz_cell.value)
        lz_cell.value = lz_str
        lz_cell.hyperlink = (
            f"https://www.bi.id.ethz.ch/orgdb/OrgeinheitDetailPre.do?leitzahl={lz_str}"
        )
        lz_cell.data_type = "s"
        lz_cell.style = "Hyperlink"

        # Create Hyperlinks for building, floor and room
        (geb, ges, raum) = row[0:3]
        sheet[
            f"A{lz_cell.row}"
        ].hyperlink = (
            f"https://ethz.ch/staffnet/de/utils/location.html?building={geb.value}"
        )
        sheet[
            f"B{lz_cell.row}"
        ].hyperlink = f"https://ethz.ch/staffnet/de/utils/location.html?building={geb.value}&floor={ges.value}"
        sheet[
            f"C{lz_cell.row}"
        ].hyperlink = f"https://ethz.ch/staffnet/de/utils/location.html?building={geb.value}&floor={ges.value}&room={raum.value}"
        sheet[f"A{lz_cell.row}"].style = "Hyperlink"
        sheet[f"B{lz_cell.row}"].style = "Hyperlink"
        sheet[f"C{lz_cell.row}"].style = "Hyperlink"

        # Sum up total area per institute and room type
        rtype = row[4].value
        inst_rtype_area[inst_lz][rtype] += row[7].value * 0.01 * row[8].value
        rtypes.append(rtype)
    raum_wb.save(filename=raum_wb_copy)


def create_summary(fte_wb, summary_filepath):
    """Read FTE workbook and add up all FTEs per institute.
    Create a new workbook with Insituts-Leitzahl, Instituts-Name and all room types
    as titles
    Every row contains: LZ, Name Institut, Buero, Aufenthaltsraum, ...
    For every room type, the area is calculated per institute."""
    sheet = fte_wb["export_employees"]

    for row in sheet.iter_rows(min_row=FTE_DATA_STARTS_AT_ROW):
        lz = str(row[0].value)
        lz = "0" + lz[1:]

        # Bisheriges Jahr ytd
        fte = row[7].value
        try:
            institute_fte[lz_institute[lz]["oe_leitzahl_oben"]] += fte
        except KeyError:
            get_institute_for_lz(lz)
            institute_fte[lz_institute[lz]["oe_leitzahl_oben"]] += fte

    # create summary workbook
    rtypes_sorted = list(set(rtypes))
    rtypes_sorted.sort()
    rtype2column = {}
    for i, v in enumerate(rtypes_sorted):
        rtype2column[v] = i

    # Start workbook and set titles
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Institut LZ", "Institut Name", "FTE", *rtypes_sorted])
    for cell in sheet["1"]:
        cell.style = title_style

    def styled_cell_for_number(ws, number):
        """Only show one digit after the comma"""
        c = Cell(ws, value=number)
        c.number_format = "0.0"
        return c

    # For every institute and room type we show the assigned area
    for instlz in sorted(inst_rtype_area.keys()):
        rtype_area = inst_rtype_area[instlz]
        area_per_rtype = [
            styled_cell_for_number(sheet, rtype_area.get(rtype, 0))
            for rtype in rtypes_sorted
        ]
        sheet.append(
            [
                instlz,
                instlz_name[instlz],
                styled_cell_for_number(sheet, institute_fte.get(instlz, 0)),
                *area_per_rtype,
            ]
        )

    workbook.save(summary_filepath)


def get_institute_for_lz(lz):
    """For a given LZ, we fetch its upper Leitzahl via webservice, to get the department.
    All information will be cached in lz_institute
    """
    # prefixes of 0's to ensure a length of 5 chars
    lz = str(lz)
    lz = "0" * (5 - len(lz)) + lz
    if lz in lz_institute:
        return lz_institute[lz]

    url = f"https://idn.ethz.ch/sap_info/lzorg?leitzahl={lz}"
    resp = requests.get(url, timeout=120)
    if not resp.ok:
        raise ClickException(
            f"Cannot fetch infos for Leitzahl: {resp.status_code} {url}"
        )
    info = resp.json()[0]
    lz_institute[lz] = info
    instlz_name[str(info["oe_leitzahl_oben"])] = info["oe_name_oben"]
    return info


if __name__ == "__main__":
    cli()
